# A ideia deste programa é exibir os arquivos existentes no diretório
# Você terá as opções: adicionar linhas, criar um arquivo ou sair do programa 

import openpyxl
import os

os.system("cls" if os.name == "nt" else "clear") # limpa a tela antes de exibir os resultados

def listar_arquivos(diretorio):
    # os.listdir(diretorio) - > Essa função do módulo os retorna uma lista com todos os arquivos e subdiretórios no diretório especificado por diretorio.
    # for f in os.listdir(diretorio) -> Este é o loop que percorre cada item (arquivo ou pasta) retornado por os.listdir(diretorio). Aqui, f representa cada nome de arquivo (ou subdiretório) na lista.
    # if f.endswith(".xlsx") -> O método .endswith(".xlsx") verifica se o nome do arquivo termina com a extensão .xlsx (ou seja, se é um arquivo Excel). Apenas os itens que atendem a essa condição são incluídos na nova lista.
    # Este é o formato da list comprehension:
    # [<resultado> for <item> in <iterável> if <condição>]
    # Ele cria uma nova lista contendo apenas os elementos que atendem à condição especificada.
    # [f for f in os.listdir(diretorio) if f.endswith(".xlsx")]
    # Para cada f em os.listdir(diretorio), adicione f à nova lista somente se f terminar com .xlsx.
    arquivos = [f for f in os.listdir(diretorio) if f.endswith(".xlsx")]
    return arquivos

def exibir_sheets(arquivo):
    book = openpyxl.load_workbook(arquivo)
    print(f"Planlhas disponiveis no arquivo {arquivo}: ")
    for i, sheet in enumerate(book.sheetnames, 1): # A função enumerate é usada para iterar sobre uma coleção e, ao mesmo tempo, gerar um índice associado a cada item.
        print(f"{i}. {sheet}")
    return book
        
def adicionar_novas_linhas(sheet):
    cabecalhos = [cell.value for cell in sheet[1]]  # Cabeçalhos da primeira linha
    qtde_linhas = int(input("Quantas novas linhas deseja adicionar? "))
    ultima_linha = sheet.max_row  # Última linha preenchida
    print(f"ATENÇÃO!!! Serão adicionadas {qtde_linhas} linhas a partir da linha {ultima_linha + 1}.")

    for k in range(1, qtde_linhas + 1):
        page_dados = []
        for j, coluna in enumerate(cabecalhos):
            entrada_dados = input(f"Linha {k}/{qtde_linhas} -> {coluna}: ")
            page_dados.append(entrada_dados)
        sheet.append(page_dados)
    print("\nNovas linhas adicionadas com sucesso!")
    
def excluir_arquivo(diretorio):
    """Exclui um arquivo escolhido pelo usuário."""
    arquivos = listar_arquivos(diretorio)
    if arquivos:
        # print("Arquivos encontrados no diretório:")
        # for i, arquivo in enumerate(arquivos, 1):
        #     print(f"{i}. {arquivo}")

        try:
            escolha = int(input("Escolha o número do arquivo que deseja excluir: ")) - 1
            if 0 <= escolha < len(arquivos):
                arquivo_escolhido = os.path.join(diretorio, arquivos[escolha])
                os.remove(arquivo_escolhido)
                print(f"Arquivo '{arquivos[escolha]}' excluído com sucesso!")
            else:
                print("Escolha inválida. Nenhum arquivo foi excluído.")
        except (ValueError, IndexError):
            print("Erro na escolha. Nenhum arquivo foi excluído.")
    else:
        print("Nenhum arquivo disponível no diretório.")

# Configurações iniciais
diretorio = "./"
extensao_file = ".xlsx"

# Listar arquivos no diretório
arquivos = listar_arquivos(diretorio)
if arquivos:
    print("Arquivos encontrados no diretório:")
    for i, arquivo in enumerate(arquivos, 1):
        print(f"{i}. {arquivo}")
else:
    print("Nenhum arquivo encontrado no diretório.")

# Escolha entre adicionar linhas ou criar um arquivo
pergunta = int(input("\nDeseja adicionar linhas (1), criar (2) um arquivo, excluir (3) um arquivo ou sair (0) do sistema? "))

book = openpyxl.Workbook()

if pergunta == 1:
    print(f"Opção escolhida {pergunta}, adicionar linhas.")
    # Abrir um arquivo existente
    arquivos = listar_arquivos(diretorio)
    if arquivos:
        try:
            escolha_arquivo = int(input("Escolha o número do arquivo para abrir: ")) - 1
            if 0 <= escolha_arquivo < len(arquivos):
                arquivo_escolhido = arquivos[escolha_arquivo]
                book = exibir_sheets(arquivo_escolhido)
                escolha_sheet = int(input("Escolha o número da planilha: ")) - 1
                sheet_escolhida = book.sheetnames[escolha_sheet]
                sheet = book[sheet_escolhida]

                # Adicionar novas linhas
                adicionar_novas_linhas(sheet)
                book.save(arquivo_escolhido)
                print(f"Alterações salvas no arquivo {arquivo_escolhido}!")
            else:
                print("Escolha inválida.")
        except (ValueError, IndexError):
            print("Erro na escolha. Reinicie o programa e tente novamente.")
    else:
        print("Não há arquivos no diretório.")

elif pergunta == 2:
    print(f"Opção Escolhida {pergunta}, criar novo arquivo.")
    # Criar um novo arquivo
    book = openpyxl.Workbook()
    nome_file = input("\nDigite o nome do arquivo (não precisa inserir a extensão '.xlsx'): ")
    nome_sheet = input("Digite o nome da planilha: ")
    book.create_sheet(nome_sheet)
    if "Sheet" in book.sheetnames:
        book.remove(book["Sheet"])

    # Criar as colunas
    cabecalhos = []
    qtde_colunas = int(input("Digite quantas colunas o arquivo terá: "))
    for i in range(1, qtde_colunas + 1):
        nome_coluna = input(f"Digite o nome da coluna {i}/{qtde_colunas}: ")
        cabecalhos.append(nome_coluna)

    page = book[nome_sheet]
    page.append(cabecalhos)

    # Inserir dados na planilha
    qtde_linhas = int(input("Digite quantas linhas terá a planilha: "))
    for k in range(1, qtde_linhas + 1):
        page_dados = []
        for j in range(qtde_colunas):
            entrada_dados = input(f"Linha {k}/{qtde_linhas} -> {cabecalhos[j]}: ")
            page_dados.append(entrada_dados)
        page.append(page_dados)

    # Salvar arquivo
    book.save(nome_file + extensao_file)
    print(f"O arquivo {nome_file + extensao_file} foi criado com sucesso!")
elif pergunta == 3:
    excluir_pergunta = '''
Você realmente deseja excluir o arquivo? Não será possível recuperar novamente!
1. sim
2. não
Digite a opção: '''
    excluir = input(excluir_pergunta)
    
    if excluir == "1":
        excluir_arquivo(diretorio)
    elif excluir == "2":   
        print("Encerrando as opções de exclusão sem excluir arquivo(s).")
    
elif pergunta == 0:
    print(f"Opção escolhida sair. Ecerrando o sistema!")
else:
    print("Escolha inválida. Reinicie o programa.")
